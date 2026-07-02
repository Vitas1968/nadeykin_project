from __future__ import annotations

import argparse
import datetime as dt
import zipfile
import xml.etree.ElementTree as ET
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover - fallback for environments without openpyxl
    load_workbook = None

_SUPPORTED_EXTENSIONS = {".xlsx"}
_WHITESPACE_RE = re.compile(r"\s+")


@dataclass(slots=True)
class RowData:
    values: list[str]
    refs: list[str]
    excel_row_index: int


@dataclass(slots=True)
class _FallbackCell:
    coordinate: str
    value: Any


class _FallbackSheet:
    def __init__(self, *, name: str, sheet_state: str = "visible", rows: list[list[_FallbackCell]] | None = None) -> None:
        self.title = name
        self.sheet_state = sheet_state
        self._rows = rows or []

    def iter_rows(self, values_only: bool = False):
        for row in self._rows:
            if values_only:
                yield tuple(cell.value for cell in row)
            else:
                yield tuple(row)


class _FallbackWorkbook:
    def __init__(self, sheets: list[_FallbackSheet]) -> None:
        self._sheets = sheets
        self.sheetnames = [sheet.title for sheet in sheets]
        self._sheets_by_name = {sheet.title: sheet for sheet in sheets}

    def __getitem__(self, sheet_name: str) -> _FallbackSheet:
        return self._sheets_by_name[sheet_name]

    def close(self) -> None:
        return


def _normalize_text(value: str) -> str:
    return _WHITESPACE_RE.sub(" ", value).strip()


def _ensure_xlsx_path(path: str | Path) -> Path:
    xlsx_path = Path(path)
    if xlsx_path.suffix.lower() not in _SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported file extension for XLSX reader: {xlsx_path.suffix or '<none>'}")
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX file does not exist: {xlsx_path}")
    return xlsx_path


def _format_float(value: float) -> str:
    if value.is_integer():
        formatted = str(int(value))
    else:
        formatted = format(value, ".15g")
    if formatted == "-0":
        formatted = "0"
    return formatted


def _normalize_cell_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = _normalize_text(value)
        return text or None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return _format_float(value)
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.time):
        return value.strftime("%H:%M:%S")
    text = _normalize_text(str(value))
    return text or None


def _sheet_is_visible(sheet) -> bool:
    return getattr(sheet, "sheet_state", "visible") == "visible"


def _iter_row_data(sheet) -> list[RowData]:
    rows: list[RowData] = []
    for excel_row_index, row in enumerate(sheet.iter_rows(values_only=False), start=1):
        cell_values: list[str] = []
        cell_refs: list[str] = []
        for cell in row:
            value = _normalize_cell_value(cell.value)
            if value is None:
                continue
            cell_values.append(value)
            cell_refs.append(cell.coordinate)
        if cell_values:
            rows.append(RowData(values=cell_values, refs=cell_refs, excel_row_index=excel_row_index))
    return rows


def _build_row_block(
    *,
    source_path: str,
    sheet_name: str,
    sheet_index: int,
    row_index: int,
    excel_row_index: int,
    values: list[str],
    refs: list[str],
    section: str | None,
) -> dict[str, Any]:
    return {
        "source_path": source_path,
        "source_type": "xlsx",
        "block_id": f"s{sheet_index:06d}_r{row_index:06d}",
        "block_type": "table_row",
        "section": section,
        "text": " | ".join(values),
        "table_index": sheet_index,
        "row_index": row_index,
        "cell_values": values,
        "sheet_name": sheet_name,
        "sheet_index": sheet_index,
        "excel_row_index": excel_row_index,
        "cell_refs": refs,
    }


def _load_workbook(path: Path):
    if load_workbook is None:
        return _load_workbook_fallback(path)
    try:
        return load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - defensive wrapper around openpyxl
        raise RuntimeError(f"Unable to open XLSX file: {path}") from exc


def _column_index_from_reference(cell_reference: str) -> int:
    match = re.match(r"^([A-Z]+)", cell_reference.upper())
    if not match:
        return 0
    column = 0
    for char in match.group(1):
        column = column * 26 + (ord(char) - ord("A") + 1)
    return column


def _fallback_cell_value(cell: ET.Element, shared_strings: list[str]) -> Any:
    cell_type = cell.get("t")
    value_element = cell.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")

    if cell_type == "inlineStr":
        inline_text_parts = []
        for text_node in cell.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"):
            if text_node.text:
                inline_text_parts.append(text_node.text)
        text = "".join(inline_text_parts).strip()
        return text or None

    if value_element is None or value_element.text is None:
        return None

    raw_value = value_element.text.strip()
    if not raw_value:
        return None

    if cell_type == "s":
        try:
            index = int(raw_value)
        except ValueError:
            return raw_value
        if 0 <= index < len(shared_strings):
            return shared_strings[index]
        return raw_value

    if cell_type == "b":
        return raw_value == "1"

    if cell_type == "str":
        return raw_value

    try:
        if any(marker in raw_value.lower() for marker in (".", "e")):
            numeric_value = float(raw_value)
            if numeric_value.is_integer():
                return int(numeric_value)
            return numeric_value
        return int(raw_value)
    except ValueError:
        return raw_value


def _read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    shared_strings_path = "xl/sharedStrings.xml"
    if shared_strings_path not in archive.namelist():
        return []

    root = ET.fromstring(archive.read(shared_strings_path))
    namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    shared_strings: list[str] = []
    for shared_string in root.findall("main:si", namespace):
        parts: list[str] = []
        for text_node in shared_string.findall(".//main:t", namespace):
            if text_node.text:
                parts.append(text_node.text)
        shared_strings.append("".join(parts))
    return shared_strings


def _read_workbook_sheets(archive: zipfile.ZipFile) -> list[tuple[str, str, str]]:
    workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_xml = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    workbook_ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    rels_ns = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}

    rel_targets: dict[str, str] = {}
    for relationship in rels_xml.findall("rel:Relationship", rels_ns):
        rel_id = relationship.get("Id")
        target = relationship.get("Target")
        if rel_id and target:
            rel_targets[rel_id] = target.lstrip("/")

    sheets: list[tuple[str, str, str]] = []
    for sheet in workbook_xml.findall("main:sheets/main:sheet", workbook_ns):
        sheet_name = sheet.get("name") or "Sheet"
        rel_id = sheet.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        target = rel_targets.get(rel_id or "", "")
        sheet_state = sheet.get("state") or "visible"
        if target:
            sheets.append((sheet_name, sheet_state, f"xl/{target}" if not target.startswith("xl/") else target))
    return sheets


def _read_xlsx_fallback(path: Path) -> _FallbackWorkbook:
    if not path.exists():
        raise FileNotFoundError(f"XLSX file does not exist: {path}")
    if path.suffix.lower() not in _SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported file extension for XLSX reader: {path.suffix or '<none>'}")

    try:
        archive = zipfile.ZipFile(path)
    except Exception as exc:  # pragma: no cover - defensive wrapper around zipfile
        raise RuntimeError(f"Unable to open XLSX file: {path}") from exc

    with archive:
        shared_strings = _read_shared_strings(archive)
        workbook_sheets = _read_workbook_sheets(archive)
        namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        sheets: list[_FallbackSheet] = []

        for sheet_name, sheet_state, sheet_path in workbook_sheets:
            try:
                sheet_root = ET.fromstring(archive.read(sheet_path))
            except Exception as exc:  # pragma: no cover - defensive wrapper around XML parsing
                raise RuntimeError(f"Unable to read XLSX sheet: {sheet_path}") from exc

            rows_by_index: dict[int, list[_FallbackCell]] = {}
            for row_element in sheet_root.findall("main:sheetData/main:row", namespace):
                row_reference = row_element.get("r")
                try:
                    row_index = int(row_reference) if row_reference else len(rows_by_index) + 1
                except ValueError:
                    row_index = len(rows_by_index) + 1

                cells: list[_FallbackCell] = []
                for cell_element in row_element.findall("main:c", namespace):
                    coordinate = cell_element.get("r") or ""
                    value = _fallback_cell_value(cell_element, shared_strings)
                    if value is None:
                        continue
                    cells.append(_FallbackCell(coordinate=coordinate, value=value))

                if cells:
                    cells.sort(key=lambda cell: _column_index_from_reference(cell.coordinate))
                    rows_by_index[row_index] = cells

            ordered_rows = [rows_by_index[row_index] for row_index in sorted(rows_by_index)]
            sheets.append(_FallbackSheet(name=sheet_name, sheet_state=sheet_state, rows=ordered_rows))

    return _FallbackWorkbook(sheets)


def _read_xlsx_content(path: str | Path) -> tuple[list[dict[str, Any]], int]:
    xlsx_path = _ensure_xlsx_path(path)
    workbook = _load_workbook(xlsx_path)
    source_path = str(xlsx_path.resolve())

    blocks: list[dict[str, Any]] = []
    visible_sheet_count = 0

    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if not _sheet_is_visible(sheet):
                continue

            visible_sheet_count += 1
            section: str | None = None
            row_index = 0

            for row_data in _iter_row_data(sheet):
                row_index += 1
                if len(row_data.values) == 1 and len(row_data.values[0]) <= 120:
                    section = row_data.values[0]
                blocks.append(
                    _build_row_block(
                        source_path=source_path,
                        sheet_name=sheet_name,
                        sheet_index=visible_sheet_count,
                        row_index=row_index,
                        excel_row_index=row_data.excel_row_index,
                        values=row_data.values,
                        refs=row_data.refs,
                        section=section,
                    )
                )
    except Exception as exc:  # pragma: no cover - defensive wrapper around openpyxl/worksheet iteration
        raise RuntimeError(f"Unable to read XLSX file: {xlsx_path}") from exc
    finally:
        workbook.close()

    return blocks, visible_sheet_count


def extract_xlsx_blocks(path: str | Path) -> list[dict[str, Any]]:
    blocks, _ = _read_xlsx_content(path)
    return blocks


def read_xlsx(path: str | Path) -> dict[str, Any]:
    blocks, visible_sheets = _read_xlsx_content(path)
    source_path = blocks[0]["source_path"] if blocks else str(Path(path).resolve())
    return {
        "source_path": source_path,
        "source_type": "xlsx",
        "blocks": blocks,
        "full_text": "\n".join(block["text"] for block in blocks if block["text"]),
        "stats": {
            "sheets": visible_sheets,
            "rows": len(blocks),
            "blocks": len(blocks),
        },
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("path")
    args = parser.parse_args()
    result = read_xlsx(args.path)
    print("blocks:", result["stats"]["blocks"])
    print("sheets:", result["stats"]["sheets"])
    print("rows:", result["stats"]["rows"])
    print(result["full_text"][:1000])
