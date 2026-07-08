#!/usr/bin/env python3
"""
write_results.py — Put the check results back where the business expects them.

It runs every raw result record (from run_batch.py) through evaluate.py — the
single source of truth for the audit rules — and writes ONLY the existing audit
columns:
  AI "Фольга снята с пачки"               -> "1" если фольга НЕ снята, иначе пусто
  AJ "На Photo URL есть фото DMC"          -> "1" если DMC не распознан, иначе пусто
  AK "DMC код совпадает с кодом на пачке"  -> "1" если DMC ≠ эталон AE, иначе пусто
  AN  итоговая оценка                      -> "1" валидный | "0" не валидный
       (AN=0 если любой из AI/AJ/AK == "1", либо AF пуст, либо AE пуст)

Output modes
------------
--mode xlsx  --template ORIGINAL.xlsx --out FILLED.xlsx
    The Excel report: loads the source workbook (all sheets/formatting kept),
    writes AI/AJ/AK/AN into their existing cells, saves to a copy. Nothing else
    is touched — no new columns. Memory-heavy on a huge workbook; fine for ~500
    rows / a few MB.

--mode csv   (default)
    Compact CSV: row, AI, AJ, AK, AN. Paste back into the workbook by row number.

Review file
-----------
--review review.csv writes ONLY rows needing human eyes (evaluate marks them):
  * foil не определён / модель не уверена / ошибка,
  * AK == 1 (DMC распознан, но не совпадает — вероятное реальное расхождение),
  * AJ == 1 (DMC не распознан ни на AG, ни на AF).
"""
import argparse
import csv
import json
import sys

from evaluate import evaluate_row


def load_results(path):
    out = {}
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rec = json.loads(line)
            out[rec["row"]] = rec
    return out


def write_csv(results, out_path, foil_unknown, dmc_only=False):
    rows = sorted(results.values(), key=lambda r: r["row"])
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        if dmc_only:
            w.writerow(["row", "AJ (нет DMC)", "AK (DMC не совпадает)"])
            for r in rows:
                ev = evaluate_row(r, foil_unknown=foil_unknown)
                w.writerow([r["row"], ev["AJ"], ev["AK"]])
        else:
            w.writerow(["row", "AI (фольга не снята)", "AJ (нет DMC)",
                        "AK (DMC не совпадает)", "AN (итог)"])
            for r in rows:
                ev = evaluate_row(r, foil_unknown=foil_unknown)
                w.writerow([r["row"], ev["AI"], ev["AJ"], ev["AK"], ev["AN"]])
    sys.stderr.write(f"[write_results] wrote {len(rows)} rows -> {out_path}\n")


def write_review(results, out_path, foil_unknown, dmc_only=False):
    rows = sorted(results.values(), key=lambda r: r["row"])
    flagged = []
    for r in rows:
        ev = evaluate_row(r, foil_unknown=foil_unknown)
        if not ev["review_needed"]:
            continue
        if ev["AK"] == "1":
            reason = "DMC распознан, но НЕ совпадает с эталоном AE"
        elif ev["AJ"] == "1":
            reason = "DMC не распознан (ни на AG, ни на AF)"
        else:
            reason = "Фольга: модель не уверена / не определено / ошибка"
        flagged.append((r, ev, reason))
    src = lambda ev: {"AG": "AG", "AF": "AF (фоллбэк)"}.get(ev["decode_source"], "—")
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        if dmc_only:
            # foil wasn't run — omit AI/AN/foil columns (they'd be misleading)
            w.writerow(["row", "Причина проверки", "Распознанный DMC", "Эталон AE",
                        "Источник DMC", "AK"])
            for r, ev, reason in flagged:
                w.writerow([r["row"], reason, ev["decoded"], ev["dmc_ref"], src(ev), ev["AK"]])
        else:
            w.writerow(["row", "Причина проверки", "Распознанный DMC", "Эталон AE",
                        "Источник DMC", "AK", "AI", "AN", "Статус фольги", "Foil confidence"])
            for r, ev, reason in flagged:
                w.writerow([
                    r["row"], reason, ev["decoded"], ev["dmc_ref"], src(ev),
                    ev["AK"], ev["AI"], ev["AN"], ev["foil_status"],
                    ev["foil_confidence"] if ev["foil_confidence"] is not None else "",
                ])
    sys.stderr.write(f"[write_results] review queue: {len(flagged)} rows -> {out_path}\n")


def write_xlsx(results, template, out_path, foil_unknown, sheet="Detail",
               col_ai="AI", col_aj="AJ", col_ak="AK", col_an="AN", dmc_only=False):
    import openpyxl
    sys.stderr.write("[write_results] loading workbook (may be slow/memory-heavy on big files)...\n")
    wb = openpyxl.load_workbook(template)
    ws = wb[sheet]
    n = 0
    for row, rec in results.items():
        ev = evaluate_row(rec, foil_unknown=foil_unknown)
        ws[f"{col_aj}{row}"] = ev["AJ"]
        ws[f"{col_ak}{row}"] = ev["AK"]
        # DMC-only run: foil wasn't checked, so AI/AN would be meaningless
        # (AI=1 / AN=0 for every row). Leave those existing cells untouched.
        if not dmc_only:
            ws[f"{col_ai}{row}"] = ev["AI"]
            ws[f"{col_an}{row}"] = ev["AN"]
        n += 1
    wb.save(out_path)
    extra = " (только AJ/AK — DMC-only)" if dmc_only else ""
    sys.stderr.write(f"[write_results] wrote {n} rows into {out_path}{extra}\n")


def summarize(results, foil_unknown, dmc_only=False):
    """Print an aggregate report + WARN on pathological distributions.

    A degenerate result (0% decoded, every download failing, AN=0 for all) means
    the PIPELINE is broken (wrong columns / images unreachable), NOT that every
    pack is invalid. Surfacing it turns a confusing "100 невалидных" into an
    actionable "картинки не качаются / не та колонка". In --dmc-only mode AN/foil
    were never evaluated, so those stats are omitted (else they'd read AN=0 100%)."""
    recs = list(results.values())
    n = len(recs)
    if not n:
        return
    an1 = an0 = decoded = dl_err = empty_url = af_absent = foil_unknown_n = 0
    for r in recs:
        ev = evaluate_row(r, foil_unknown=foil_unknown)
        an1 += ev["AN"] == "1"
        an0 += ev["AN"] == "0"
        decoded += bool(ev["dmc_found"])
        if not r.get("af_present"):
            af_absent += 1
        derr = (r.get("decode_error") or "")
        if "download_error" in derr:
            dl_err += 1
        elif derr == "empty_url":
            empty_url += 1
        if r.get("AI_foil_removed") is None and r.get("foil_checked"):
            foil_unknown_n += 1

    pct = lambda x: f"{x}/{n} ({100*x//n}%)"
    if dmc_only:
        sys.stderr.write(
            f"[write_results] СВОДКА (DMC-only, AN/фольга не оценивались): строк={n}; "
            f"DMC распознан {pct(decoded)}; ошибка загрузки фото {pct(dl_err)}; "
            f"пустой URL {pct(empty_url)}; AF без ссылки {pct(af_absent)}\n")
    else:
        sys.stderr.write(
            f"[write_results] СВОДКА: строк={n}; AN=1 {pct(an1)}; AN=0 {pct(an0)}; "
            f"DMC распознан {pct(decoded)}; foil не определён {pct(foil_unknown_n)}; "
            f"ошибка загрузки фото {pct(dl_err)}; пустой URL {pct(empty_url)}; AF без ссылки {pct(af_absent)}\n"
        )
    if decoded == 0:
        sys.stderr.write(
            "[write_results] WARNING: DMC не распознан НИ В ОДНОЙ строке. Это сбой пайплайна, "
            "а не «вся партия невалидна»: либо не та колонка/лист в extract, либо картинки "
            "недоступны (сеть/VPN/хост). Проверь decode_error в results.jsonl ПЕРЕД выводами.\n")
    if dl_err >= n * 0.9 and n:
        sys.stderr.write(
            "[write_results] WARNING: >90% строк — ошибка загрузки фото. Хост картинок недоступен "
            "со стенда (сеть/VPN/файрвол), это НЕ баг extract и НЕ плохие данные.\n")
    if empty_url >= n * 0.9 and n:
        sys.stderr.write(
            "[write_results] WARNING: >90% строк — пустой URL. extract отдал пустые ссылки: "
            "не та буква колонки или не тот лист. Сверь заголовок и --col-*/--sheet.\n")


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("results", help="results.jsonl from run_batch.py")
    ap.add_argument("--mode", choices=["csv", "xlsx"], default="csv")
    ap.add_argument("--out", default="results.csv", help="Output file")
    ap.add_argument("--review", default=None, help="Also write a review-queue CSV here")
    ap.add_argument("--foil-unknown", choices=["problem", "ok"], default="problem",
                    help="How to treat a foil status the model could not determine. "
                         "problem (default): AI=1 and the row is not valid (AN=0). "
                         "ok: AI stays empty.")
    ap.add_argument("--template", default=None, help="(xlsx mode) source .xlsx to copy + fill")
    ap.add_argument("--sheet", default="Detail", help="(xlsx mode) worksheet name")
    ap.add_argument("--dmc-only", action="store_true",
                    help="DMC-only run (no --foil in run_batch): fill ONLY AJ/AK, leave "
                         "AI and AN untouched (otherwise foil-less AI=1/AN=0 would be wrong).")
    args = ap.parse_args()

    results = load_results(args.results)
    if not results:
        sys.exit("[write_results] no results to write")

    if args.mode == "csv":
        write_csv(results, args.out, args.foil_unknown, dmc_only=args.dmc_only)
    else:
        if not args.template:
            ap.error("--mode xlsx requires --template ORIGINAL.xlsx")
        write_xlsx(results, args.template, args.out, args.foil_unknown,
                   sheet=args.sheet, dmc_only=args.dmc_only)

    if args.review:
        write_review(results, args.review, args.foil_unknown, dmc_only=args.dmc_only)

    summarize(results, args.foil_unknown, dmc_only=args.dmc_only)


if __name__ == "__main__":
    main()
